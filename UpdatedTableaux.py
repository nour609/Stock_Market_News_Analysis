import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from io import BytesIO
import re
from datetime import datetime

# Données simulées des titres d’actualités
data = {
    "Date": pd.date_range(start="2020-01-01", periods=10, freq="D"),
    "Label": [1, 0, 1, 1, 0, 1, 0, 1, 0, 1],
    "Top1": ["Stock market gains as investors cheer news"] * 10,
    "Top2": ["Economy shows signs of recovery despite global uncertainty"] * 10,
    "Top3": ["Tech companies lead rally amid strong earnings reports"] * 10,
}
df = pd.DataFrame(data)

# Données historiques des réactions du marché
market_reactions = pd.DataFrame([
    {"Date": "2025-10-10", "Événement": "Selloff due to trade tensions", "Impact": "DJIA ↓ 1.90%"},
    {"Date": "2016-06-24", "Événement": "Brexit referendum result", "Impact": "DJIA ↓ 611 pts"},
    {"Date": "2008-09-15", "Événement": "Lehman Brothers collapse", "Impact": "DJIA ↓ 500 pts"},
    {"Date": "2020-03-16", "Événement": "COVID-19 lockdown announcement", "Impact": "DJIA ↓ 2,997 pts"},
    {"Date": "2011-08-08", "Événement": "US credit rating downgrade", "Impact": "DJIA ↓ 634.76 pts"},
    {"Date": "1987-10-19", "Événement": "Black Monday crash", "Impact": "DJIA ↓ 22.6%"},
    {"Date": "2022-02-24", "Événement": "Russia invades Ukraine", "Impact": "DJIA ↓ 800 pts"},
    {"Date": "2020-11-09", "Événement": "COVID-19 vaccine announcement", "Impact": "DJIA ↑ 834 pts"},
    {"Date": "2001-09-11", "Événement": "9/11 terrorist attacks", "Impact": "DJIA ↓ 684.81 pts"},
    {"Date": "1997-10-27", "Événement": "Asian financial crisis", "Impact": "DJIA ↓ 554.26 pts"},
    {"Date": "1989-10-13", "Événement": "UAL leveraged buyout failure", "Impact": "DJIA ↓ 190.58 pts"},
    {"Date": "1973-10-06", "Événement": "OPEC oil embargo", "Impact": "DJIA ↓ 45.79 pts"},
    {"Date": "1962-05-28", "Événement": "Cuban Missile Crisis", "Impact": "DJIA ↓ 5.7%"},
    {"Date": "1994-02-23", "Événement": "Mexican peso crisis", "Impact": "DJIA ↓ 2.2%"},
    {"Date": "2000-03-10", "Événement": "Dot-com bubble burst", "Impact": "NASDAQ ↓ 78%"},
    {"Date": "2015-08-24", "Événement": "China devalues yuan", "Impact": "DJIA ↓ 588.40 pts"},
    {"Date": "2018-02-05", "Événement": "Volatility spike", "Impact": "DJIA ↓ 1,175 pts"},
    {"Date": "2019-08-14", "Événement": "Inverted yield curve", "Impact": "DJIA ↓ 800 pts"},
    {"Date": "2021-01-06", "Événement": "Capitol riot", "Impact": "DJIA ↓ 400 pts"},
    {"Date": "2022-03-16", "Événement": "Fed raises interest rates", "Impact": "DJIA ↓ 500 pts"},
    {"Date": "2014-07-17", "Événement": "MH17 shot down", "Impact": "DJIA ↓ 161.39 pts"},
    {"Date": "2017-08-17", "Événement": "Charlottesville unrest", "Impact": "DJIA ↓ 274.14 pts"},
    {"Date": "2023-11-09", "Événement": "Midterm elections result", "Impact": "DJIA ↑ 400 pts"},
    {"Date": "2024-05-05", "Événement": "Fed signals rate hike pause", "Impact": "DJIA ↑ 350 pts"},
    {"Date": "2025-02-20", "Événement": "Tech earnings beat expectations", "Impact": "NASDAQ ↑ 2.5%"},
    {"Date": "2025-07-30", "Événement": "China stimulus announced", "Impact": "DJIA ↑ 600 pts"},
], columns=["Date", "Événement", "Impact"])

# Convertir les dates en datetime
market_reactions["Date"] = pd.to_datetime(market_reactions["Date"])

# Extraire la valeur numérique et signe de l'impact
def parse_impact(impact_str):
    # Extrait le signe ↑ ou ↓
    sign = 1
    if "↓" in impact_str:
        sign = -1
    # Extrait le nombre (enlever virgule dans milliers)
    num_str = re.findall(r"[\d,.]+", impact_str)[-1]
    num_str = num_str.replace(",", "")
    try:
        value = float(num_str)
    except:
        value = 0.0
    return sign * value

market_reactions["Impact_Num"] = market_reactions["Impact"].apply(parse_impact)

# Création du fichier Excel
wb = Workbook()

# Feuille 1 : Données des titres d’actualités
ws1 = wb.active
ws1.title = "Stock News Sample"
ws1.append(df.columns.tolist())
for row in df.itertuples(index=False):
    ws1.append(row)

# Feuille 2 : Réactions du marché
ws2 = wb.create_sheet(title="Market Reactions")
ws2.append(market_reactions.columns.tolist())
for row in market_reactions.itertuples(index=False):
    ws2.append(row)

# Création du graphique
plt.figure(figsize=(12, 6))
plt.bar(market_reactions["Date"], market_reactions["Impact_Num"],
        color=market_reactions["Impact_Num"].apply(lambda x: 'green' if x > 0 else 'red'))
plt.axhline(0, color='black', linewidth=0.8)
plt.title("Impact des événements majeurs sur le marché (points ou %)")
plt.ylabel("Variation (positive = hausse, negative = baisse)")
plt.xlabel("Date")
plt.xticks(rotation=45)
plt.tight_layout()

# Enregistrer le graphique en mémoire
img_data = BytesIO()
plt.savefig(img_data, format='png')
plt.close()
img_data.seek(0)

# Insérer le graphique dans la feuille Excel
img = Image(img_data)
img.width = 800
img.height = 400
ws2.add_image(img, "G2")  # Positionne le graphique à partir de la cellule G2

# Enregistrer le fichier Excel
filename = "Stock_News_and_Market_Reactions_25events.xlsx"
wb.save(filename)

print(f"Fichier Excel généré : {filename}")
