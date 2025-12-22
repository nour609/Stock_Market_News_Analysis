import pandas as pd
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment

# -------- Données simulées des titres d’actualités -------- #
data = {
    "Date": pd.date_range(start="2020-01-01", periods=10, freq="D"),
    "Label": [random.choice([0, 1]) for _ in range(10)],
    "Top1": ["Stock market gains as investors cheer news"] * 10,
    "Top2": ["Economy shows signs of recovery despite global uncertainty"] * 10,
    "Top3": ["Tech companies lead rally amid strong earnings reports"] * 10,
}
df = pd.DataFrame(data)

# -------- Données historiques réelles des réactions du marché -------- #
market_reactions = pd.DataFrame([
    {
        "Date": "2008-09-15",
        "Événement/Nouvelle": "Faillite de Lehman Brothers",
        "Type d'événement": "Crise financière",
        "Impact sur le marché": "DJIA ↓ 500 pts",
        "Détails": "Déclenchement de la crise financière mondiale"
    },
    {
        "Date": "2016-06-24",
        "Événement/Nouvelle": "Résultats du Brexit",
        "Type d'événement": "Événement politique",
        "Impact sur le marché": "DJIA ↓ 611 pts",
        "Détails": "Sortie surprise du Royaume-Uni de l’UE"
    },
    {
        "Date": "2020-03-16",
        "Événement/Nouvelle": "Confinement mondial COVID-19",
        "Type d'événement": "Crise sanitaire",
        "Impact sur le marché": "DJIA ↓ 2 997 pts",
        "Détails": "Pire jour du Dow Jones depuis 1987"
    },
    {
        "Date": "2022-02-24",
        "Événement/Nouvelle": "Invasion de l’Ukraine par la Russie",
        "Type d'événement": "Conflit géopolitique",
        "Impact sur le marché": "DJIA ↓ 800 pts (puis rebond)",
        "Détails": "Volatilité extrême dans la journée"
    },
    {
        "Date": "2020-11-09",
        "Événement/Nouvelle": "Annonce d’un vaccin COVID",
        "Type d'événement": "Bonne nouvelle",
        "Impact sur le marché": "DJIA ↑ 834 pts",
        "Détails": "Optimisme économique mondial"
    },
])

# -------- Création du fichier Excel -------- #
wb = Workbook()

# Feuille 1 : Données de titres
ws1 = wb.active
ws1.title = "Stock News Sample"
ws1.append(df.columns.tolist())
for cell in ws1[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

for i, row in df.iterrows():
    ws1.append(list(row))
    comment = Comment(f"Date: {row['Date'].strftime('%Y-%m-%d')}\nLabel: {'Up/Stable' if row['Label'] == 1 else 'Down'}", "AI")
    ws1.cell(row=i+2, column=1).comment = comment

table1 = Table(displayName="NewsData", ref=f"A1:D{len(df)+1}")
style1 = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
table1.tableStyleInfo = style1
ws1.add_table(table1)

# Feuille 2 : Réactions du marché
ws2 = wb.create_sheet(title="Market Reactions")
ws2.append(market_reactions.columns.tolist())
for cell in ws2[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")

for row in market_reactions.itertuples(index=False):
    ws2.append(list(row))

table2 = Table(displayName="MarketEvents", ref=f"A1:E{len(market_reactions)+1}")
style2 = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
table2.tableStyleInfo = style2
ws2.add_table(table2)

# Enregistrement du fichier
excel_filename = "Stock_News_and_Reactions.xlsx"
wb.save(excel_filename)
print(f"Fichier enregistré : {excel_filename}")
